from __future__ import annotations


import openpyxl
import pandas as pd


class TestExtractGroundTruth:
    def test_returns_dict_keyed_by_sheet_name(self, tmp_xlsx_full):
        from excel2py.verifier import extract_ground_truth

        result = extract_ground_truth(tmp_xlsx_full)
        assert "Sales" in result

    def test_data_cells_captured(self, tmp_xlsx_full):
        from excel2py.verifier import extract_ground_truth

        result = extract_ground_truth(tmp_xlsx_full)
        df = result["Sales"]
        assert df.shape[0] > 0
        assert df.shape[1] > 0

    def test_merged_cell_anchor_value_present(self, tmp_xlsx_full):
        from excel2py.verifier import extract_ground_truth

        result = extract_ground_truth(tmp_xlsx_full)
        df = result["Sales"]
        first_row_values = [str(v) for v in df.iloc[0].tolist() if v is not None]
        assert any("Sales Report" in v for v in first_row_values)

    def test_empty_sheet_excluded(self, tmp_path):
        from excel2py.verifier import extract_ground_truth

        filepath = tmp_path / "empty.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Empty"
        wb.save(filepath)
        wb.close()
        result = extract_ground_truth(filepath)
        assert "Empty" not in result

    def test_excel_error_cells_skipped(self, tmp_path):
        from excel2py.verifier import extract_ground_truth

        filepath = tmp_path / "errors.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "#REF!"
        ws["A2"] = 42
        wb.save(filepath)
        wb.close()
        result = extract_ground_truth(filepath)
        if "Sheet1" in result:
            df = result["Sheet1"]
            flat = df.values.flatten().tolist()
            assert "#REF!" not in flat


class TestCompareOutputs:
    def test_no_output_files_with_empty_ground_truth_passes(self, tmp_path):
        from excel2py.verifier import compare_outputs

        result = compare_outputs(tmp_path, {})
        assert result.passed

    def test_no_output_files_with_ground_truth_fails(self, tmp_path):
        from excel2py.verifier import compare_outputs

        ground_truth = {"Sales": pd.DataFrame([[1, 2], [3, 4]])}
        result = compare_outputs(tmp_path, ground_truth)
        assert not result.passed
        assert any(e.error_type == "no_output" for e in result.errors)

    def test_matching_csv_passes(self, tmp_path):
        from excel2py.verifier import compare_outputs

        csv_file = tmp_path / "Sales.csv"
        csv_file.write_text("Widget,10.0\nGadget,25.0\n")
        ground_truth = {"Sales": pd.DataFrame([["Widget", 10.0], ["Gadget", 25.0]])}
        result = compare_outputs(tmp_path, ground_truth)
        assert result.passed

    def test_value_mismatch_fails(self, tmp_path):
        from excel2py.verifier import compare_outputs

        csv_file = tmp_path / "Sales.csv"
        csv_file.write_text("Widget,99.0\nGadget,25.0\n")
        ground_truth = {"Sales": pd.DataFrame([["Widget", 10.0], ["Gadget", 25.0]])}
        result = compare_outputs(tmp_path, ground_truth)
        assert not result.passed
        assert any(e.error_type == "mismatch" for e in result.errors)

    def test_missing_sheet_fails(self, tmp_path):
        from excel2py.verifier import compare_outputs

        ground_truth = {"Sales": pd.DataFrame([[1, 2]])}
        result = compare_outputs(tmp_path, ground_truth)
        assert not result.passed
        assert any(
            e.error_type in ("no_output", "missing_sheet") for e in result.errors
        )

    def test_numeric_tolerance(self, tmp_path):
        from excel2py.verifier import compare_outputs

        csv_file = tmp_path / "Sheet1.csv"
        csv_file.write_text("1.000000001\n")
        ground_truth = {"Sheet1": pd.DataFrame([[1.0]])}
        result = compare_outputs(tmp_path, ground_truth)
        assert result.passed


class TestRunScript:
    def test_successful_script_returns_zero(self, tmp_path):
        from excel2py.verifier import run_script

        script = tmp_path / "ok.py"
        script.write_text("import sys\nprint('ok')\n")
        xlsx = tmp_path / "dummy.xlsx"
        xlsx.write_bytes(b"")
        exit_code, stdout, stderr, output_dir = run_script(script, xlsx, timeout=10)
        assert exit_code == 0
        assert "ok" in stdout

    def test_crashing_script_returns_nonzero(self, tmp_path):
        from excel2py.verifier import run_script

        script = tmp_path / "crash.py"
        script.write_text("raise RuntimeError('boom')\n")
        xlsx = tmp_path / "dummy.xlsx"
        xlsx.write_bytes(b"")
        exit_code, stdout, stderr, output_dir = run_script(script, xlsx, timeout=10)
        assert exit_code != 0
        assert "boom" in stderr

    def test_timeout_returns_minus_one(self, tmp_path):
        from excel2py.verifier import run_script

        script = tmp_path / "hang.py"
        script.write_text("import time\ntime.sleep(999)\n")
        xlsx = tmp_path / "dummy.xlsx"
        xlsx.write_bytes(b"")
        exit_code, stdout, stderr, output_dir = run_script(script, xlsx, timeout=1)
        assert exit_code == -1
        assert "timed out" in stderr.lower()

    def test_output_files_written_to_output_dir(self, tmp_path):
        from excel2py.verifier import run_script

        script = tmp_path / "writer.py"
        script.write_text("open('out.csv', 'w').write('a,b\\n1,2\\n')\n")
        xlsx = tmp_path / "dummy.xlsx"
        xlsx.write_bytes(b"")
        exit_code, stdout, stderr, output_dir = run_script(script, xlsx, timeout=10)
        assert exit_code == 0
        assert (output_dir / "out.csv").exists()


class TestBuildCorrectionPrompt:
    def test_includes_original_code(self):
        from excel2py.verifier import VerificationResult, build_correction_prompt

        result = VerificationResult(passed=False, errors=[])
        prompt = build_correction_prompt("print('hello')", result, 0, "")
        assert "print('hello')" in prompt

    def test_includes_crash_stderr(self):
        from excel2py.verifier import VerificationResult, build_correction_prompt

        result = VerificationResult(passed=False, errors=[])
        prompt = build_correction_prompt("x", result, 1, "NameError: name 'x'")
        assert "NameError" in prompt

    def test_includes_mismatch_table(self):
        from excel2py.verifier import (
            VerificationError,
            VerificationResult,
            build_correction_prompt,
        )

        errors = [VerificationError("Sales", "row 0, col 1", 10.0, 99.0, "mismatch")]
        result = VerificationResult(passed=False, errors=errors)
        prompt = build_correction_prompt("code", result, 0, "")
        assert "Sales" in prompt
        assert "10.0" in prompt
        assert "99.0" in prompt


class TestLintGeneratedCode:
    def test_to_csv_missing_header_false(self):
        from excel2py.verifier import lint_generated_code
        code = "import pandas as pd\ndf.to_csv('out.csv', index=False)\n"
        findings = lint_generated_code(code)
        assert any("header=False" in f for f in findings)

    def test_to_csv_missing_index_false(self):
        from excel2py.verifier import lint_generated_code
        code = "import pandas as pd\ndf.to_csv('out.csv', header=False)\n"
        findings = lint_generated_code(code)
        assert any("index=False" in f for f in findings)

    def test_to_csv_both_flags_present_no_finding(self):
        from excel2py.verifier import lint_generated_code
        code = "import pandas as pd\ndf.to_csv('out.csv', index=False, header=False)\n"
        findings = lint_generated_code(code)
        assert not any("to_csv" in f for f in findings)

    def test_if_v_in_items_loop_detected(self):
        from excel2py.verifier import lint_generated_code
        code = (
            "for (r, c), v in values.items():\n"
            "    if v:\n"
            "        data[r][c] = v\n"
        )
        findings = lint_generated_code(code)
        assert any("if v:" in f for f in findings)

    def test_if_v_outside_loop_not_flagged(self):
        from excel2py.verifier import lint_generated_code
        code = "if v:\n    do_something()\n"
        findings = lint_generated_code(code)
        assert not any("if v:" in f for f in findings)

    def test_dropna_without_how_detected(self):
        from excel2py.verifier import lint_generated_code
        code = "df = df.dropna()\n"
        findings = lint_generated_code(code)
        assert any("dropna" in f for f in findings)

    def test_dropna_with_how_all_no_finding(self):
        from excel2py.verifier import lint_generated_code
        code = "df = df.dropna(how='all')\n"
        findings = lint_generated_code(code)
        assert not any("dropna" in f for f in findings)

    def test_syntax_error_returns_empty(self):
        from excel2py.verifier import lint_generated_code
        findings = lint_generated_code("def (broken code")
        assert findings == []

    def test_findings_injected_into_correction_prompt(self):
        from excel2py.verifier import VerificationResult, build_correction_prompt
        result = VerificationResult(passed=False, errors=[])
        prompt = build_correction_prompt(
            "code", result, 0, "",
            lint_findings=["Line 5: to_csv() is missing header=False."],
        )
        assert "Static Analysis Findings" in prompt
        assert "header=False" in prompt


class TestClassifyMismatchErrors:
    def test_nan_vs_empty_majority(self):
        from excel2py.verifier import VerificationError, _classify_mismatch_errors

        errors = [
            VerificationError("Draw", f"row {i}, col 1", "", float("nan"), "mismatch")
            for i in range(6)
        ]
        assert _classify_mismatch_errors(errors) == "nan_vs_empty"

    def test_merged_cell_when_no_empty_string(self):
        from excel2py.verifier import VerificationError, _classify_mismatch_errors

        errors = [
            VerificationError("Draw", f"row {i}, col 1", "Team A", None, "mismatch")
            for i in range(4)
        ]
        assert _classify_mismatch_errors(errors) == "merged_cell"

    def test_empty_errors_returns_merged_cell(self):
        from excel2py.verifier import _classify_mismatch_errors

        assert _classify_mismatch_errors([]) == "merged_cell"

    def test_mixed_majority_nan_vs_empty(self):
        from excel2py.verifier import VerificationError, _classify_mismatch_errors

        nan_empty = [
            VerificationError("S", f"row {i}", "", float("nan"), "mismatch")
            for i in range(7)
        ]
        other = [VerificationError("S", "row 9", "A", "B", "mismatch") for _ in range(3)]
        assert _classify_mismatch_errors(nan_empty + other) == "nan_vs_empty"


class TestBuildValueMismatchLastResort:
    def test_nan_vs_empty_path_contains_fillna(self):
        from excel2py.verifier import VerificationError, VerificationResult, build_value_mismatch_last_resort_prompt

        errors = [
            VerificationError("Draw", f"row {i}, col 1", "", float("nan"), "mismatch")
            for i in range(6)
        ]
        result = VerificationResult(passed=False, errors=errors)
        prompt = build_value_mismatch_last_resort_prompt("code", {}, result)
        assert "fillna" in prompt
        assert "is not None" in prompt
        assert "openpyxl" not in prompt  # must NOT inject merged-cell pattern

    def test_merged_cell_path_contains_openpyxl(self):
        from excel2py.verifier import VerificationError, VerificationResult, build_value_mismatch_last_resort_prompt

        errors = [
            VerificationError("Draw", f"row {i}, col 1", "Team A", None, "mismatch")
            for i in range(4)
        ]
        result = VerificationResult(passed=False, errors=errors)
        prompt = build_value_mismatch_last_resort_prompt("code", {}, result)
        assert "openpyxl" in prompt
        assert "merged_fill" in prompt


class TestSystemPromptRule10:
    def test_system_prompt_contains_empty_string_rule(self):
        from excel2py.prompts.templates import get_system_prompt
        prompt = get_system_prompt()
        assert "fillna" in prompt
        assert "is not None" in prompt


class TestPlateauExit:
    def test_plateau_exit_breaks_before_max_attempts(self, tmp_path):
        """Converter exits early when best_score stagnates for PLATEAU_WINDOW attempts."""
        from unittest.mock import MagicMock, patch
        import pandas as pd
        from excel2py.converter import convert
        from excel2py.verifier import VerificationResult, VerificationError

        dummy_code = "import sys\nprint('hello')"
        stale_errors = [
            VerificationError("Sheet1", "row 0, col 0", "", float("nan"), "mismatch")
            for _ in range(5)
        ]
        stale_result = VerificationResult(passed=False, errors=stale_errors)

        call_count = {"n": 0}

        def fake_compare(output_dir, ground_truth):
            call_count["n"] += 1
            return stale_result

        dummy_excel = tmp_path / "test.xlsx"
        import openpyxl
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        wb.active["A1"] = "hello"
        wb.save(dummy_excel)

        ground_truth = {"Sheet1": pd.DataFrame([["hello"]])}

        with (
            patch("excel2py.converter.create_chat_model") as mock_lm_factory,
            patch("excel2py.converter.extract_ground_truth", return_value=ground_truth),
            patch("excel2py.converter.compare_outputs", side_effect=fake_compare),
            patch("excel2py.converter.run_script", return_value=(0, "", "", tmp_path)),
            patch("excel2py.converter.check_properties", return_value=[]),
            patch("excel2py.converter.compute_output_diagnostics", return_value=None),
            patch("excel2py.converter.run_output_judge", return_value=""),
            patch("excel2py.converter.run_rubber_duck_diagnosis", return_value="ROOT CAUSE: nan"),
            patch("excel2py.converter.run_langchain_correction", return_value=dummy_code),
        ):
            mock_lm = MagicMock()
            mock_lm.invoke.return_value = MagicMock(content=dummy_code)
            mock_lm_factory.return_value = mock_lm

            convert(
                tmp_path / "test.xlsx",
                provider="openai",
                api_key="fake",
                model="gpt-4o",
                max_verify_attempts=5,
                verify=True,
            )

        # With plateau_window=2 and no improvement, should exit before attempt 5
        assert call_count["n"] < 5
