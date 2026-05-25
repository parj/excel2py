import pytest
import openpyxl


@pytest.fixture
def tmp_xlsx(tmp_path):
    """Create a simple .xlsx file with formulas and data."""
    filepath = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    # Headers
    ws["A1"] = "Product"
    ws["B1"] = "Price"
    ws["C1"] = "Quantity"
    ws["D1"] = "Total"

    # Data
    ws["A2"] = "Widget"
    ws["B2"] = 10.0
    ws["C2"] = 5
    ws["D2"] = "=B2*C2"

    ws["A3"] = "Gadget"
    ws["B3"] = 25.0
    ws["C3"] = 3
    ws["D3"] = "=B3*C3"

    # Summary formula
    ws["D4"] = "=SUM(D2:D3)"

    wb.save(filepath)
    wb.close()
    return filepath


@pytest.fixture
def tmp_xlsx_with_named_range(tmp_path):
    """Create .xlsx with a named range."""
    filepath = tmp_path / "named.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = 100
    ws["A2"] = 200
    from openpyxl.workbook.defined_name import DefinedName

    dn = DefinedName("MyRange", attr_text="'Data'!$A$1:$A$2")
    wb.defined_names.add(dn)
    wb.save(filepath)
    wb.close()
    return filepath


@pytest.fixture
def tmp_xlsx_merged(tmp_path):
    """Create .xlsx with merged cells."""
    filepath = tmp_path / "merged.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Merged Header"
    ws.merge_cells("A1:C1")
    ws["A2"] = 1
    ws["B2"] = 2
    ws["C2"] = 3
    wb.save(filepath)
    wb.close()
    return filepath


@pytest.fixture
def tmp_xlsx_full(tmp_path):
    """Excel with data, formulas, and merged cells for verifier tests."""
    filepath = tmp_path / "full.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    # Row 1: merged header
    ws["A1"] = "Sales Report"
    ws.merge_cells("A1:C1")

    # Row 2: column headers
    ws["A2"] = "Product"
    ws["B2"] = "Price"
    ws["C2"] = "Total"

    # Row 3-4: data + formula
    ws["A3"] = "Widget"
    ws["B3"] = 10.0
    ws["C3"] = "=B3*2"

    ws["A4"] = "Gadget"
    ws["B4"] = 25.0
    ws["C4"] = "=B4*2"

    wb.save(filepath)
    wb.close()
    return filepath
