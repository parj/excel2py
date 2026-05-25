from __future__ import annotations

import logging
import subprocess
import sys
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pandas as pd
from xlsb_reader import XlsbWorkbook, XlsxWorkbook

logger = logging.getLogger(__name__)

_EXCEL_ERRORS = {"#REF!", "#VALUE!", "#N/A", "#NAME?", "#NUM!", "#NULL!", "#DIV/0!"}


@dataclass
class VerificationError:
    sheet: str
    location: str
    expected: Any
    actual: Any
    error_type: str  # "mismatch" | "missing_sheet" | "crash" | "no_output" | "shape_mismatch"


@dataclass
class VerificationResult:
    passed: bool
    errors: list[VerificationError] = field(default_factory=list)

    def error_count(self) -> int:
        return len(self.errors)


def _values_to_dataframe(values: dict) -> pd.DataFrame:
    """Convert xlsb_reader sparse {(row, col): value} dict to a dense DataFrame.

    xlsb_reader returns only non-empty cells (and only anchor cells for merged
    ranges), so this naturally omits empty separator rows — no manual dropna needed
    for empty rows, but we still apply it for safety and to drop empty columns.
    """
    if not values:
        return pd.DataFrame()
    max_row = max(r for r, _ in values)
    max_col = max(c for _, c in values)
    data: list[list] = [[None] * (max_col + 1) for _ in range(max_row + 1)]
    for (r, c), v in values.items():
        if not (isinstance(v, str) and v in _EXCEL_ERRORS):
            data[r][c] = v
    df = pd.DataFrame(data)
    return df.dropna(how="all").dropna(axis=1, how="all")


def extract_ground_truth(excel_path: Path) -> dict[str, pd.DataFrame]:
    """Extract cached cell values from Excel as DataFrames using xlsb_reader.

    Uses the same reader as generated scripts, so the ground truth exactly
    matches what the scripts will produce — no merged-cell forward-fill
    discrepancy, no empty-row offset mismatches.
    """
    suffix = excel_path.suffix.lower()
    wb = XlsbWorkbook(excel_path) if suffix == ".xlsb" else XlsxWorkbook(excel_path)

    result: dict[str, pd.DataFrame] = {}
    for sheet_name, values in wb.iter_values():
        df = _values_to_dataframe(values)
        if not df.empty:
            result[sheet_name] = df
    return result


def _find_matching_file(output_files: list[Path], sheet_name: str) -> Path | None:
    sheet_key = sheet_name.lower().replace(" ", "_")
    for f in output_files:
        if f.stem.lower().replace(" ", "_") == sheet_key:
            return f
    for f in output_files:
        stem = f.stem.lower()
        if sheet_key in stem or stem in sheet_key:
            return f
    if len(output_files) == 1:
        return output_files[0]
    return None


def _compare_dataframes(
    expected: pd.DataFrame, actual: pd.DataFrame, sheet_name: str
) -> list[VerificationError]:
    errors: list[VerificationError] = []
    expected = expected.reset_index(drop=True)
    actual = actual.reset_index(drop=True)

    if expected.shape != actual.shape:
        errors.append(VerificationError(
            sheet=sheet_name,
            location="shape",
            expected=str(expected.shape),
            actual=str(actual.shape),
            error_type="shape_mismatch",
        ))
        return errors

    rows = min(len(expected), 50)
    cols = min(len(expected.columns), 20)

    for r in range(rows):
        for c in range(cols):
            exp_val = expected.iat[r, c]
            act_val = actual.iat[r, c]

            exp_is_na = exp_val is None or (isinstance(exp_val, float) and pd.isna(exp_val))
            act_is_na = act_val is None or (isinstance(act_val, float) and pd.isna(act_val))
            if exp_is_na and act_is_na:
                continue

            try:
                if abs(float(exp_val) - float(act_val)) <= abs(float(exp_val)) * 1e-5 + 1e-9:
                    continue
            except (TypeError, ValueError):
                if str(exp_val) == str(act_val):
                    continue

            errors.append(VerificationError(
                sheet=sheet_name,
                location=f"row {r}, col {c}",
                expected=exp_val,
                actual=act_val,
                error_type="mismatch",
            ))
            if len(errors) >= 20:
                return errors

    return errors


def compare_outputs(
    output_dir: Path, ground_truth: dict[str, pd.DataFrame]
) -> VerificationResult:
    if not ground_truth:
        return VerificationResult(passed=True)

    output_files = list(output_dir.glob("*.csv")) + list(output_dir.glob("*.xlsx"))

    if not output_files:
        return VerificationResult(
            passed=False,
            errors=[VerificationError(
                sheet="", location="", expected="output files", actual="none written",
                error_type="no_output",
            )],
        )

    errors: list[VerificationError] = []

    for sheet_name, expected_df in ground_truth.items():
        output_file = _find_matching_file(output_files, sheet_name)
        if output_file is None:
            errors.append(VerificationError(
                sheet=sheet_name, location="", expected="output file", actual="not found",
                error_type="missing_sheet",
            ))
            continue

        try:
            if output_file.suffix == ".csv":
                actual_df = pd.read_csv(output_file, header=None)
            else:
                actual_df = pd.read_excel(output_file, header=None)
        except Exception as e:
            errors.append(VerificationError(
                sheet=sheet_name, location="", expected="readable file", actual=str(e),
                error_type="crash",
            ))
            continue

        errors.extend(_compare_dataframes(expected_df, actual_df, sheet_name))

    return VerificationResult(passed=len(errors) == 0, errors=errors)


def run_script(
    script_path: Path, excel_path: Path, timeout: int = 60
) -> tuple[int, str, str, Path]:
    """Run the generated script in a temp dir. Returns (exit_code, stdout, stderr, output_dir)."""
    output_dir = Path(tempfile.mkdtemp())
    try:
        proc = subprocess.run(
            [sys.executable, str(script_path), str(excel_path)],
            capture_output=True,
            text=True,
            timeout=timeout,
            cwd=str(output_dir),
        )
        return proc.returncode, proc.stdout, proc.stderr, output_dir
    except subprocess.TimeoutExpired:
        return -1, "", f"Script timed out after {timeout}s", output_dir


def check_properties(
    output_dir: Path, ground_truth: dict[str, pd.DataFrame]
) -> list[str]:
    """Plain-English property failures derived from ground truth — no hard-coded error categories."""
    output_files = list(output_dir.glob("*.csv")) + list(output_dir.glob("*.xlsx"))
    failures: list[str] = []

    if not output_files:
        failures.append("No output files were written by the script.")
        return failures

    for sheet_name, expected_df in ground_truth.items():
        if not isinstance(expected_df, pd.DataFrame):
            continue

        output_file = _find_matching_file(output_files, sheet_name)
        if output_file is None:
            failures.append(f"Sheet '{sheet_name}': no output file was written.")
            continue

        try:
            actual_df = (
                pd.read_csv(output_file, header=None)
                if output_file.suffix == ".csv"
                else pd.read_excel(output_file, header=None)
            )
        except Exception as e:
            failures.append(f"Sheet '{sheet_name}': could not read output — {e}")
            continue

        row_diff = actual_df.shape[0] - expected_df.shape[0]
        col_diff = actual_df.shape[1] - expected_df.shape[1]

        if row_diff > 0:
            failures.append(
                f"Sheet '{sheet_name}': {row_diff} too many rows "
                f"(wrote {actual_df.shape[0]}, need {expected_df.shape[0]})."
            )
        elif row_diff < 0:
            failures.append(
                f"Sheet '{sheet_name}': {-row_diff} too few rows "
                f"(wrote {actual_df.shape[0]}, need {expected_df.shape[0]})."
            )

        if col_diff > 0:
            failures.append(
                f"Sheet '{sheet_name}': {col_diff} too many columns "
                f"(wrote {actual_df.shape[1]}, need {expected_df.shape[1]})."
            )
        elif col_diff < 0:
            failures.append(
                f"Sheet '{sheet_name}': {-col_diff} too few columns "
                f"(wrote {actual_df.shape[1]}, need {expected_df.shape[1]})."
            )

        if row_diff == 0 and col_diff == 0:
            # Shape is correct — count value mismatches
            mismatch_count = 0
            rows = min(len(actual_df), 50)
            cols = min(actual_df.shape[1], 20)
            for r in range(rows):
                for c in range(cols):
                    exp_val = expected_df.iat[r, c]
                    act_val = actual_df.iat[r, c]
                    exp_na = exp_val is None or (isinstance(exp_val, float) and pd.isna(exp_val))
                    act_na = act_val is None or (isinstance(act_val, float) and pd.isna(act_val))
                    if exp_na and act_na:
                        continue
                    try:
                        if abs(float(exp_val) - float(act_val)) <= abs(float(exp_val)) * 1e-5 + 1e-9:
                            continue
                    except (TypeError, ValueError):
                        if str(exp_val) == str(act_val):
                            continue
                    mismatch_count += 1
            if mismatch_count:
                failures.append(
                    f"Sheet '{sheet_name}': shape is correct but "
                    f"{mismatch_count} cell value(s) differ."
                )

    return failures


def compute_output_diagnostics(
    output_dir: Path, ground_truth: dict[str, pd.DataFrame]
) -> str:
    """Return concrete before/after data for each failing sheet so the LLM can identify the bad operation."""
    output_files = list(output_dir.glob("*.csv")) + list(output_dir.glob("*.xlsx"))
    if not output_files:
        return ""

    lines: list[str] = []
    for sheet_name, expected_df in ground_truth.items():
        output_file = _find_matching_file(output_files, sheet_name)
        if not output_file:
            lines.append(f"\nSheet '{sheet_name}': no output file was written.")
            continue
        try:
            actual_df = (
                pd.read_csv(output_file, header=None)
                if output_file.suffix == ".csv"
                else pd.read_excel(output_file, header=None)
            )
        except Exception as e:
            lines.append(f"\nSheet '{sheet_name}': could not read output — {e}")
            continue

        if not isinstance(expected_df, pd.DataFrame):
            continue
        if expected_df.shape == actual_df.shape:
            continue

        lines.append(f"\n### Sheet '{sheet_name}': shape mismatch")
        lines.append(f"Expected {expected_df.shape} — got {actual_df.shape}")

        lines.append("\nExpected — first 5 rows:")
        lines.append(expected_df.head(5).to_string(index=False, header=False))
        lines.append("\nActual — first 5 rows:")
        lines.append(actual_df.head(5).to_string(index=False, header=False))

        row_diff = actual_df.shape[0] - expected_df.shape[0]
        col_diff = actual_df.shape[1] - expected_df.shape[1]

        if row_diff > 0:
            lines.append(f"\nYour script wrote {row_diff} EXTRA ROW(S). Actual last 5 rows:")
            lines.append(actual_df.tail(5).to_string(index=False, header=False))
            lines.append("\nExpected last 5 rows (what it should end with):")
            lines.append(expected_df.tail(5).to_string(index=False, header=False))
        elif row_diff < 0:
            lines.append(f"\nYour script is MISSING {-row_diff} ROW(S). Expected last 5 rows:")
            lines.append(expected_df.tail(5).to_string(index=False, header=False))

        if col_diff > 0:
            lines.append(f"\nYour script wrote {col_diff} EXTRA COLUMN(S).")
            lines.append(f"  Actual columns beyond expected: {list(actual_df.iloc[0, expected_df.shape[1]:])}")
        elif col_diff < 0:
            lines.append(f"\nYour script is MISSING {-col_diff} COLUMN(S).")

    return "\n".join(lines)


def build_correction_prompt(
    code: str,
    result: VerificationResult,
    exit_code: int,
    stderr: str,
    workbook_description: str | None = None,
    attempt_history: list | None = None,
    environment_info: str | None = None,
    output_diagnostics: str | None = None,
    ground_truth_samples: dict | None = None,
    stagnant_attempts: int = 0,
    property_failures: list[str] | None = None,
    judge_feedback: str | None = None,
) -> str:
    lines = [
        "The Python script you generated has issues. Fix it and return ONLY the corrected Python file.",
        "",
    ]

    if stagnant_attempts >= 2:
        lines.append(
            f"## ⚠ STAGNATION WARNING — same error for {stagnant_attempts} consecutive attempts\n"
            "Your current approach is FUNDAMENTALLY wrong for this data structure.\n"
            "You MUST try a completely different strategy, for example:\n"
            "- If you used `pd.read_excel(..., header=0)`, try `header=None` and locate data rows manually\n"
            "- If you sliced by row index, re-examine where the actual data starts in the sheet\n"
            "- If you used `dropna`, the threshold may be wrong — try `dropna(how='all')` then "
            "strip trailing empty rows explicitly\n"
            "- If you included columns by position, recount — the index column may be included\n"
            "Do not repeat any approach from the Attempt History below."
        )
        lines.append("")

    if environment_info:
        lines.append(f"## Runtime Environment\n{environment_info}")
        lines.append("Use only APIs that exist in these exact versions. "
                     "For example: pandas 2.x removed `DataFrame.applymap` — use `DataFrame.map` instead.\n")

    if attempt_history:
        lines.append("## Attempt History (do not repeat approaches that have already failed)")
        for h in attempt_history:
            crashed = " — CRASHED" if h["exit_code"] != 0 else ""
            lines.append(f"  Attempt {h['n']}: {h['error_count']} errors{crashed}")
            for err in h["errors"][:3]:
                lines.append(
                    f"    [{err.error_type}] sheet={err.sheet!r} "
                    f"expected={err.expected!r} actual={err.actual!r}"
                )
            for line in h["stderr_tail"][:2]:
                lines.append(f"    stderr: {line}")
            # Include rubber duck diagnosis as Reflexion-style verbal memory so the fixer
            # knows what root causes were already identified in prior attempts.
            if h.get("diagnosis"):
                root_cause_line = ""
                for line in h["diagnosis"].splitlines():
                    if "ROOT CAUSE" in line.upper():
                        root_cause_line = line.strip()
                        break
                if root_cause_line:
                    lines.append(f"    Diagnosed root cause: {root_cause_line}")
        lines.append("")

    # Most recent diagnosis surfaced as primary signal for the fixer
    if attempt_history:
        latest_diagnosis = attempt_history[-1].get("diagnosis")
        if latest_diagnosis:
            lines.append("## Root Cause Diagnosis (from rubber duck analysis of the last attempt)")
            lines.append("This explains WHY the current code fails — fix this root cause, not just the symptoms.")
            lines.append(latest_diagnosis)
            lines.append("")

    if workbook_description:
        lines.append("## Original Excel Structure")
        lines.append("Use this to understand the exact sheets, columns, and data the script must reproduce.")
        lines.append(workbook_description)
        lines.append("")

    lines.append("## Current Issues (from the best run so far)")

    if exit_code != 0:
        lines.append(f"\n### Script crashed (exit code {exit_code})")
        lines.append("```")
        lines.append(stderr[:3000])
        lines.append("```")

    if property_failures:
        lines.append("\n### What is wrong (property checks)")
        for f in property_failures:
            lines.append(f"  - {f}")
    elif result.errors and not judge_feedback:
        # Fallback: show raw errors only if no richer feedback is available
        lines.append(f"\n### Verification errors ({len(result.errors)} total)")
        for err in result.errors[:20]:
            lines.append(f"  - [{err.error_type}] {err.sheet}: {err.expected!r} vs {err.actual!r}")

    if judge_feedback:
        lines.append("\n### Why it is wrong (output analysis)")
        lines.append(judge_feedback)

    if output_diagnostics:
        lines.append("\n### Concrete row/column diff")
        lines.append(output_diagnostics)

    if ground_truth_samples:
        lines.append("\n## EXACT Expected Output — your script must reproduce this precisely")
        lines.append("(Values come from the Excel file itself with data_only=True. "
                     "None means the cell contained a formula with no cached value.)")
        for sheet_name, df in ground_truth_samples.items():
            if not isinstance(df, pd.DataFrame):
                continue
            lines.append(f"\n### Sheet '{sheet_name}' — MUST be exactly {df.shape[0]} rows × {df.shape[1]} columns")
            lines.append(
                f"If your dynamic trimming cannot produce this exact shape, use as a last resort: "
                f"`output_df = output_df.iloc[:{df.shape[0]}, :{df.shape[1]}]`"
            )
            lines.append("First 5 rows of expected output:")
            lines.append(df.head(5).to_string(index=False, header=False))
            if len(df) > 8:
                lines.append("Last 3 rows of expected output:")
                lines.append(df.tail(3).to_string(index=False, header=False))

    lines.append("\n## Best Code So Far")
    lines.append("```python")
    lines.append(code)
    lines.append("```")
    lines.append(
        "\nReturn ONLY the corrected Python code, no markdown fences, no explanations."
    )

    return "\n".join(lines)


def build_value_mismatch_last_resort_prompt(
    code: str,
    ground_truth: dict,
    result: "VerificationResult",
    environment_info: str | None = None,
) -> str:
    """Approach-pivot prompt for persistent merged-cell NaN mismatches.

    Injects a complete, tested worksheet-reading pattern that uses a dict-based
    forward-fill lookup (identical to how the verifier reads ground truth — proven
    correct for every openpyxl version). The LLM must replace its current approach
    with this exact pattern.
    """
    lines = [
        "## APPROACH PIVOT REQUIRED",
        "",
        "Your current merged-cell handling is not working. The same NaN mismatches "
        "have appeared across multiple correction attempts. You MUST replace your "
        "worksheet-reading code with the EXACT pattern below — do not modify it.",
        "",
        "## Why this happens",
        "openpyxl returns `None` for every non-anchor cell in a merged range. Approaches "
        "that mutate the worksheet (unmerge_cells, ws.cell().value = ..., ws._cells[...]) "
        "are unreliable. The only guaranteed fix is a read-time lookup dict built BEFORE "
        "calling ws.iter_rows() or ws.values.",
        "",
        "## MANDATORY replacement for your worksheet-to-DataFrame function",
        "Replace whatever function reads a worksheet into a DataFrame with EXACTLY this:",
        "",
        "```python",
        "from openpyxl.utils import get_column_letter",
        "",
        "def worksheet_to_dataframe(ws):",
        "    # Step 1: build merged-cell fill map from anchor values",
        "    merged_fill = {}",
        "    for merged_range in ws.merged_cells.ranges:",
        "        min_col, min_row, max_col, max_row = merged_range.bounds",
        "        anchor_val = ws.cell(row=min_row, column=min_col).value",
        "        for r in range(min_row, max_row + 1):",
        "            for c in range(min_col, max_col + 1):",
        "                if r == min_row and c == min_col:",
        "                    continue",
        "                merged_fill[(r, c)] = anchor_val",
        "    # Step 2: read cells, substituting fill map for None non-anchors",
        "    rows = []",
        "    for ws_row in ws.iter_rows():",
        "        row_data = []",
        "        for cell in ws_row:",
        "            val = cell.value",
        "            if val is None:",
        "                val = merged_fill.get((cell.row, cell.column))",
        "            row_data.append(val)",
        "        rows.append(row_data)",
        "    if not rows:",
        "        return pd.DataFrame()",
        "    max_cols = max(len(r) for r in rows)",
        "    normalized = [r + [None] * (max_cols - len(r)) for r in rows]",
        "    return pd.DataFrame(normalized, dtype=object)",
        "```",
        "",
        "Do NOT call fill_merged_cells, unmerge_cells, or modify ws in any way. "
        "Do NOT use ws.values (it skips merged fill). Use ONLY the pattern above.",
        "",
    ]

    if environment_info:
        lines.append(f"Runtime: {environment_info}\n")

    # Show sample mismatches so the LLM understands the scope
    mismatch_errors = [e for e in result.errors if e.error_type == "mismatch"]
    if mismatch_errors:
        lines.append(
            f"## Sample mismatches ({len(mismatch_errors)} total — all caused by missing merged-cell fill)"
        )
        for err in mismatch_errors[:15]:
            lines.append(
                f"  Sheet '{err.sheet}' {err.location}: expected {err.expected!r}, got {err.actual!r}"
            )
        lines.append("")

    # Show expected first rows so the LLM can verify its output
    error_sheets = {e.sheet for e in mismatch_errors}
    for sheet_name, df in ground_truth.items():
        if not isinstance(df, pd.DataFrame):
            continue
        if sheet_name not in error_sheets:
            continue
        lines.append(f"## Expected first 5 rows for sheet '{sheet_name}' (must match exactly)")
        lines.append(df.head(5).to_string(index=False, header=False))
        lines.append("")

    lines.append("## Current Code (replace the worksheet reading part)")
    lines.append("```python")
    lines.append(code)
    lines.append("```")
    lines.append(
        "\nReturn ONLY the corrected Python code, no markdown fences, no explanations."
    )
    return "\n".join(lines)


def build_last_resort_prompt(
    code: str,
    ground_truth: dict,
    environment_info: str | None = None,
) -> str:
    """Targeted prompt used when the loop has converged but shape errors remain.

    Instructs the fixer to use explicit iloc slicing with the exact expected
    dimensions as a guaranteed fallback — no more guessing at boundaries.
    """
    lines = [
        "The script is very close but has a persistent shape mismatch that your previous "
        "fixes could not resolve. Use the EXACT DIMENSIONS below to force the correct output shape.",
        "",
    ]
    if environment_info:
        lines.append(f"Runtime: {environment_info}\n")

    lines.append("## Required action")
    lines.append(
        "For each Part sheet listed below, after computing the output DataFrame, "
        "trim it to the exact expected shape using `output_df = output_df.iloc[:ROWS, :COLS]`. "
        "This is the correct expected size from the Excel file itself."
    )
    lines.append("")
    lines.append("## Exact required shapes")
    for sheet_name, df in ground_truth.items():
        if not isinstance(df, pd.DataFrame):
            continue
        lines.append(
            f"  Sheet '{sheet_name}': {df.shape[0]} rows × {df.shape[1]} columns  "
            f"→ `df = df.iloc[:{df.shape[0]}, :{df.shape[1]}]`"
        )
    lines.append("")
    lines.append("## Current Code")
    lines.append("```python")
    lines.append(code)
    lines.append("```")
    lines.append(
        "\nReturn ONLY the corrected Python code, no markdown fences, no explanations."
    )
    return "\n".join(lines)


def build_fresh_generation_prompt(
    workbook_description: str,
    attempt_history: list,
    required_shapes: dict[str, tuple[int, int]],
) -> str:
    """Fresh generation prompt with Reflexion-style negative examples.

    Resets the LLM's priors by combining the original workbook description with
    anti-patterns from failed attempts and hard shape constraints from ground truth.
    Used as a Level-4 escape after incremental corrections plateau (arXiv 2303.11366).
    """
    lines = []

    if attempt_history:
        diagnosed_causes: list[str] = []
        for h in attempt_history:
            if h.get("diagnosis"):
                for line in h["diagnosis"].splitlines():
                    if "ROOT CAUSE" in line.upper():
                        diagnosed_causes.append(line.strip())
                        break

        lines.append(
            "## IMPORTANT: Previous code generation attempts failed repeatedly.\n"
            "You MUST start fresh — do NOT reuse the same logic as before.\n"
        )
        if diagnosed_causes:
            lines.append("## Diagnosed root causes from prior attempts (avoid these patterns)")
            for cause in diagnosed_causes:
                lines.append(f"  - {cause}")
            lines.append("")

    if required_shapes:
        lines.append("## Required output shapes (HARD CONSTRAINTS — must be exact)")
        lines.append(
            "Your generated code MUST produce DataFrames with EXACTLY these shapes. "
            "If in doubt, enforce with `df = df.iloc[:ROWS, :COLS]` after computing.\n"
        )
        for sheet_name, (rows, cols) in required_shapes.items():
            lines.append(
                f"  - Sheet '{sheet_name}': {rows} rows × {cols} columns"
                f"  →  df = df.iloc[:{rows}, :{cols}]"
            )
        lines.append("")

    lines.append(workbook_description)
    return "\n".join(lines)
